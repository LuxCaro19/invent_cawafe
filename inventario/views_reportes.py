import openpyxl
from django.http import HttpResponse
from inventario.models import Equipo
from datetime import date
from django.shortcuts import render

def vista_reportes(request):
    return render(request, 'inventario/reportes.html')

def reporte_inventario_general(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario General"

    ws.append(["Etiqueta", "Modelo", "Marca", "Ubicación", "Estado", "Asignado a"])

    for eq in Equipo.objects.select_related('modelo__marca', 'ubicacion', 'estado', 'usuario_asignado'):
        ubicacion_real = ''
        if eq.en_bodega:
            ubicacion_real = eq.ubicacion.nombre if eq.ubicacion else ''
        elif eq.usuario_asignado and eq.usuario_asignado.ubicacion:
            ubicacion_real = eq.usuario_asignado.ubicacion.nombre

        ws.append([
            eq.etiqueta,
            eq.modelo.nombre if eq.modelo else '',
            eq.modelo.marca.marca if eq.modelo and eq.modelo.marca else '',
            ubicacion_real,
            eq.estado.nombre if eq.estado else '',
            eq.usuario_asignado.nombre_completo if eq.usuario_asignado else '—'
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Inventario_General.xlsx'
    wb.save(response)
    return response

def reporte_equipos_para_venta(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipos para Venta"

    ws.append(["Etiqueta", "Modelo", "Marca", "Estado", "Ubicación", "Asignado a"])

    # Filtramos por nombre exacto de estados relacionados a venta
    estados_venta = ["Vendido", "Por Vender", "Pendiente RRHH"]
    equipos = Equipo.objects.filter(estado__nombre__in=estados_venta)

    for eq in equipos:
        ubicacion_real = ''
        if eq.en_bodega:
            ubicacion_real = eq.ubicacion.nombre if eq.ubicacion else ''
        elif eq.usuario_asignado and eq.usuario_asignado.ubicacion:
            ubicacion_real = eq.usuario_asignado.ubicacion.nombre

        ws.append([
            eq.etiqueta,
            eq.modelo.nombre if eq.modelo else '',
            eq.modelo.marca.marca if eq.modelo and eq.modelo.marca else '',
            eq.estado.nombre if eq.estado else '',
            ubicacion_real,
            eq.usuario_asignado.nombre_completo if eq.usuario_asignado else '—'
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Equipos_para_Venta.xlsx'
    wb.save(response)
    return response


def reporte_antiguedad_equipos(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Antigüedad de Equipos"

    ws.append(["Etiqueta", "Modelo", "Marca", "Fecha Ingreso", "Antigüedad (días)", "Estado", "Ubicación"])

    equipos = Equipo.objects.exclude(fecha_de_compra__isnull=True)

    for eq in equipos:
        fecha_referencia = eq.fecha_de_compra
        antiguedad = (date.today() - fecha_referencia).days

        ubicacion_real = ''
        if eq.en_bodega:
            ubicacion_real = eq.ubicacion.nombre if eq.ubicacion else ''
        elif eq.usuario_asignado and eq.usuario_asignado.ubicacion:
            ubicacion_real = eq.usuario_asignado.ubicacion.nombre

        ws.append([
            eq.etiqueta,
            eq.modelo.nombre if eq.modelo else '',
            eq.modelo.marca.marca if eq.modelo and eq.modelo.marca else '',
            fecha_referencia.strftime('%Y-%m-%d'),
            antiguedad,
            eq.estado.nombre if eq.estado else '',
            ubicacion_real
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Antiguedad_Equipos.xlsx'
    wb.save(response)
    return response
